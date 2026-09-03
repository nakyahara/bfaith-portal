# -*- coding: utf-8 -*-
"""STA パックリストExcel (輸送箱の梱包情報) の構造検出パーサ。

要件定義 (FBA納品箱詰め記録_要件定義_20260902.md §4) の確定事項:
  - 行列位置は SKU 数・箱数で変動する → ハードコード禁止、ラベル文字列から検出する
  - 未知構造は処理せず fail-closed (ok:false で返し、呼び出し側が手動転記へ誘導)
  - シート保護あり。書き込み対象 (箱数量・箱数・重量寸法) が locked なら書けない前提で報告する

出力: stdout に JSON 1個。ok:false でも exit 0 (呼び出し側は JSON の ok を見る)。
JSON を出せない致命異常だけ exit 2。
"""
import json
import sys
import hashlib
import re

MAX_SHEETS = 20          # packing sheet の上限 (異常ファイルの暴走防止)
MAX_SKU_ROWS = 2000
MAX_BOXES = 200

LABEL_GROUP = '梱包グループ'
LABEL_TOTAL_BOXES = '輸送箱の合計数'
LABEL_BOX_NAME = '輸送箱の名前'
LABEL_WEIGHT = '輸送箱の重量'
LABEL_WIDTH = '輸送箱の幅'
LABEL_LENGTH = '輸送箱の長さ'
LABEL_HEIGHT = '輸送箱の高さ'
# ヘッダ行の必須列 (見つからなければ未知形式)
REQUIRED_HEADERS = ['SKU', '商品名', 'ASIN', 'FNSKU', '予定数量', '輸送箱の数']
BOX_QTY_RE = re.compile(r'輸送箱(\d+)の数量')


def fail(error, message, **extra):
    out = {'ok': False, 'error': error, 'message': message}
    out.update(extra)
    print(json.dumps(out, ensure_ascii=False))
    sys.exit(0)


def cell_text(v):
    if v is None:
        return ''
    return str(v).strip()


def parse_sheet(ws):
    """1枚の「輸送箱の梱包情報」シートを構造検出して dict を返す。不明点は err を返す"""
    info = {
        'sheetName': ws.title,
        'protected': bool(ws.protection.sheet),
    }
    group_label = None
    group_id = None
    total_boxes_cell = None
    header_row = None
    box_name_row = None
    dim_rows = {}

    max_row = min(ws.max_row, 3000)
    max_col = min(ws.max_column, 300)

    for r in range(1, max_row + 1):
        a = cell_text(ws.cell(row=r, column=1).value)
        if a.startswith(LABEL_GROUP) and group_label is None:
            group_label = a
            # 同じ行の右側セルに pgxxxx… の梱包グループID
            for c in range(2, max_col + 1):
                v = cell_text(ws.cell(row=r, column=c).value)
                if v.startswith('pg'):
                    group_id = v
                    break
        if a == LABEL_BOX_NAME and box_name_row is None:
            box_name_row = r
        for key, label in (('weight', LABEL_WEIGHT), ('width', LABEL_WIDTH),
                           ('length', LABEL_LENGTH), ('height', LABEL_HEIGHT)):
            if a.startswith(label) and key not in dim_rows:
                dim_rows[key] = r
        if header_row is None and a == 'SKU':
            header_row = r
        if total_boxes_cell is None:
            for c in range(1, max_col + 1):
                v = cell_text(ws.cell(row=r, column=c).value)
                if v.startswith(LABEL_TOTAL_BOXES):
                    # ラベルの右側で最初に数値が入っているセル = 箱数
                    for c2 in range(c + 1, max_col + 1):
                        v2 = ws.cell(row=r, column=c2).value
                        if isinstance(v2, (int, float)):
                            total_boxes_cell = {'row': r, 'col': c2, 'value': int(v2)}
                            break
                    break

    if not group_id:
        return None, 'no_group_id'
    if header_row is None:
        return None, 'no_header_row'
    if total_boxes_cell is None:
        return None, 'no_total_boxes'
    if box_name_row is None or box_name_row <= header_row:
        return None, 'no_box_name_row'
    if len(dim_rows) != 4:
        return None, 'no_dim_rows'

    # ヘッダ列マップ (ラベル文字列 → 列番号)。空白ゆらぎは strip で吸収。
    # 箱数量列は =IF(...) の数式なので静的ラベルだけを拾う (fingerprint を箱数に依存させない)
    headers = {}
    for c in range(1, max_col + 1):
        v = cell_text(ws.cell(row=header_row, column=c).value)
        if v and not v.startswith('='):
            headers[v] = c
    for need in REQUIRED_HEADERS:
        if need not in headers:
            return None, f'missing_header:{need}'

    # 箱数量列: ヘッダ行の数式 =IF(...>=n,"輸送箱nの数量","") から箱番号→列を確定
    box_cols = {}
    for c in range(1, max_col + 1):
        cell = ws.cell(row=header_row, column=c)
        text = ''
        if isinstance(cell.value, str):
            text = cell.value
        m = BOX_QTY_RE.search(text)
        if m:
            n = int(m.group(1))
            if n in box_cols:
                return None, 'duplicate_box_col'
            box_cols[n] = c
    if not box_cols:
        return None, 'no_box_columns'
    # Amazon 側の箱名 (=IF(M3>=n,"P1 - Bn","") の文字列部分)。表示・箱札用で、無ければ None
    box_names = {}
    for n, c in box_cols.items():
        v = ws.cell(row=box_name_row, column=c).value
        m = re.search(r'"([^"]+)"', v) if isinstance(v, str) and v.startswith('=') else None
        box_names[str(n)] = m.group(1) if m else None
    ns = sorted(box_cols)
    if ns[0] != 1 or ns != list(range(1, len(ns) + 1)):
        return None, 'box_columns_not_contiguous'
    if len(ns) > MAX_BOXES:
        return None, 'too_many_box_columns'

    # SKU 行: ヘッダ行の下〜箱名行の上。SKU か FNSKU が入っている行だけ (予備の空行は飛ばす)
    sku_rows = []
    for r in range(header_row + 1, box_name_row):
        sku = cell_text(ws.cell(row=r, column=headers['SKU']).value)
        fnsku = cell_text(ws.cell(row=r, column=headers['FNSKU']).value)
        if not sku and not fnsku:
            continue
        qty_raw = ws.cell(row=r, column=headers['予定数量']).value
        if not isinstance(qty_raw, (int, float)):
            return None, f'bad_planned_qty:row{r}'
        row = {
            'row': r,
            'sku': sku,
            'productName': cell_text(ws.cell(row=r, column=headers['商品名']).value),
            'excelId': cell_text(ws.cell(row=r, column=headers.get('Id', headers['SKU'])).value) if 'Id' in headers else '',
            'asin': cell_text(ws.cell(row=r, column=headers['ASIN']).value),
            'fnsku': fnsku,
            'plannedQty': int(qty_raw),
        }
        sku_rows.append(row)
        if len(sku_rows) > MAX_SKU_ROWS:
            return None, 'too_many_sku_rows'
    if not sku_rows:
        return None, 'no_sku_rows'

    # 書き込み対象セルの locked 検査 (シート保護時に書けないセルへ書く事故の予防線)。
    # 対象 = 箱数セル・SKU行×箱列・寸法重量行×箱列。
    # ⚠実物テンプレの箱数量セルは「空セル + 列スタイル(unlocked)」— openpyxl は空セルに
    #   既定(locked=True)を返すので、明示スタイルが無いセルは列スタイルの protection を見る
    from openpyxl.utils import get_column_letter

    def col_style_locked(c):
        cd = ws.column_dimensions.get(get_column_letter(c))
        idx = getattr(cd, 'style', None) if cd is not None else None
        if idx is None:
            return True  # 列スタイルなし → 既定 locked
        try:
            style_array = ws.parent._cell_styles[int(idx)]
            prot = ws.parent._protections[style_array.protectionId]
            return bool(prot.locked) if prot.locked is not None else True
        except Exception:
            return True

    locked = []
    def check_locked(r, c, what):
        if not bool(ws.protection.sheet):
            return
        cell = ws.cell(row=r, column=c)
        is_locked = cell.protection.locked if cell.has_style else col_style_locked(c)
        if is_locked:
            locked.append({'what': what, 'row': r, 'col': c})
    check_locked(total_boxes_cell['row'], total_boxes_cell['col'], 'total_boxes')
    # 入力済みセルの検出 (Codex PR2 #2): STA からDLした直後のテンプレは箱数量・重量寸法が全て空。
    # 値が残っている = 記入済み/出力済みファイルの再アップロード → 取込側で拒否する (古い値の混入防止)
    prefilled = []
    def check_prefilled(r, c, what):
        v = ws.cell(row=r, column=c).value
        if v is not None and str(v).strip() != '':
            prefilled.append({'what': what, 'row': r, 'col': c})
    for row in sku_rows:
        for n in ns:
            check_locked(row['row'], box_cols[n], 'qty')
            check_prefilled(row['row'], box_cols[n], 'qty')
    for key, r in dim_rows.items():
        for n in ns:
            check_locked(r, box_cols[n], key)
            check_prefilled(r, box_cols[n], key)

    info.update({
        'packingGroupLabel': group_label,
        'packingGroupId': group_id,
        'totalBoxes': total_boxes_cell,
        'headerRow': header_row,
        'headers': headers,
        'boxColumns': {str(n): box_cols[n] for n in ns},
        'boxNames': box_names,
        'maxBoxColumns': len(ns),
        'boxNameRow': box_name_row,
        'dimRows': dim_rows,
        'skuRows': sku_rows,
        'lockedTargets': locked,
        'prefilledTargets': prefilled,
    })
    return info, None


class ParseError(Exception):
    """analyze() の失敗 (error, message, extra)"""
    def __init__(self, error, message, **extra):
        super().__init__(message)
        self.error = error
        self.message = message
        self.extra = extra


def analyze(path):
    """ワークブックを構造解析して結果 dict を返す。未知形式は ParseError"""
    def fail(error, message, **extra):
        raise ParseError(error, message, **extra)
    try:
        import openpyxl
    except Exception as e:  # venv 未整備
        fail('no_openpyxl', f'openpyxl が読み込めません: {e}')
    try:
        wb = openpyxl.load_workbook(path, data_only=False, read_only=False, keep_vba=False)
    except Exception as e:
        fail('bad_xlsx', f'Excelとして開けません: {e}')

    # Metadata シートは必須 (無い・読めない = 未知形式として拒否。Codex PR1 #2)
    if 'Metadata' not in wb.sheetnames:
        fail('no_metadata', 'Metadataシートがありません (STAのパックリストではない可能性)')
    meta = {}
    ws = wb['Metadata']
    for r in range(1, min(ws.max_row, 50) + 1):
        k = cell_text(ws.cell(row=r, column=1).value)
        v = ws.cell(row=r, column=2).value
        if k:
            meta[k] = v if isinstance(v, (int, float)) else cell_text(v)

    sheets = []
    errors = []
    for name in wb.sheetnames:
        if name in ('Metadata', '手順'):
            continue
        if len(sheets) >= MAX_SHEETS:
            errors.append({'sheet': name, 'error': 'too_many_sheets'})
            break
        info, err = parse_sheet(wb[name])
        if err:
            errors.append({'sheet': name, 'error': err})
        else:
            sheets.append(info)

    # 1シートでも解析できなければ全体を拒否する (Codex PR1 #2: 梱包グループの欠落した
    # 納品回を作らせない。部分成功は「成功」ではない)
    if errors:
        fail('sheet_parse_errors',
             '解析できないシートがあります: ' + ', '.join(f"{e['sheet']} ({e['error']})" for e in errors),
             metadata=meta, sheetErrors=errors,
             detected=[s['sheetName'] for s in sheets])
    if not sheets:
        fail('no_packing_sheets', '梱包情報シートを検出できませんでした (未知の形式)', metadata=meta, sheetErrors=errors)

    # Metadata の枚数と実検出数の突合 (欠け・型不正も未知形式として拒否)
    declared = meta.get('Number of packing sheets')
    if not isinstance(declared, (int, float)) or int(declared) != len(sheets):
        fail('sheet_count_mismatch',
             f'Metadataの梱包シート数 ({declared!r}) と検出数 ({len(sheets)}) が一致しません',
             metadata=meta, sheetErrors=errors,
             detected=[s['sheetName'] for s in sheets])

    # 構造 fingerprint (既知形式の判定材料。列ラベル・行構造だけで、値は含めない)
    fp_src = json.dumps({
        'version': meta.get('Version'),
        'locale': meta.get('Locale'),
        'sheets': [{
            'headers': sorted(s['headers'].keys()),
            'dimKeys': sorted(s['dimRows'].keys()),
        } for s in sheets],
    }, ensure_ascii=False, sort_keys=True)
    fingerprint = hashlib.sha256(fp_src.encode('utf-8')).hexdigest()[:16]

    return {
        'ok': True,
        'metadata': meta,
        'fingerprint': fingerprint,
        'sheets': sheets,
        'sheetErrors': errors,
    }


def main():
    if len(sys.argv) != 2:
        fail('bad_args', 'usage: parse_packlist.py <xlsx>')
    try:
        result = analyze(sys.argv[1])
    except ParseError as e:
        fail(e.error, e.message, **e.extra)
    print(json.dumps(result, ensure_ascii=False))


if __name__ == '__main__':
    try:
        main()
    except SystemExit:
        raise
    except Exception as e:
        print(json.dumps({'ok': False, 'error': 'internal', 'message': str(e)}, ensure_ascii=False))
        sys.exit(2)
